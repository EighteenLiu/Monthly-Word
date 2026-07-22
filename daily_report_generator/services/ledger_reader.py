from __future__ import annotations

import shutil
import uuid
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl
import xlrd
from openpyxl.utils import get_column_letter
from PIL import Image

from .models import LedgerImage, LedgerRecord
from .normalizer import (
    PHOTO_GROUP_NAMES,
    SUPPORTED_TYPES,
    clean_cell,
    normalize_station_type,
    parse_date_like,
    parse_datetime_like,
    report_date_from_created_time,
)


@dataclass
class ParsedLedger:
    records: list[LedgerRecord]
    available_dates: dict[date, int]
    warnings: list[str]
    image_count: int
    source_path: Path


def read_ledger_dates(path: Path) -> dict[date, int]:
    """Fast path for date selection: read cells only, never convert or extract images."""
    if not path.exists():
        raise FileNotFoundError(f"台账文件不存在：{path}")
    if path.suffix.lower() == ".xlsx":
        rows = read_xlsx_rows(path)
    elif path.suffix.lower() == ".xls":
        rows = read_xls_rows(path)
    else:
        raise ValueError("台账仅支持 .xls 或 .xlsx 文件")
    if len(rows) < 2:
        raise ValueError("台账至少需要两行表头")

    group_headers = fill_merged_headers([clean_cell(value) for value in rows[0]])
    raw_headers = [clean_cell(value) for value in rows[1]]
    headers = [header or group_headers[index] for index, header in enumerate(raw_headers)]
    field_indexes = build_field_indexes(group_headers, headers)
    point_type_index = field_indexes.get("point_type")
    date_index = field_indexes.get("created_time")
    if date_index is None:
        date_index = field_indexes.get("report_time")
    if date_index is None:
        raise ValueError("台账缺少日期字段：创建时间或案件上报时间")

    available_dates: dict[date, int] = {}
    for row in rows[2:]:
        point_type_value = row[int(point_type_index)] if point_type_index is not None and int(point_type_index) < len(row) else ""
        if normalize_station_type(clean_cell(point_type_value)) not in SUPPORTED_TYPES:
            continue
        date_value = row[int(date_index)] if date_index is not None and int(date_index) < len(row) else ""
        parsed_datetime = parse_datetime_like(date_value)
        parsed_date = report_date_from_created_time(parsed_datetime)
        if parsed_date:
            available_dates[parsed_date] = available_dates.get(parsed_date, 0) + 1
    return dict(sorted(available_dates.items()))


def read_ledger(path: Path, work_dir: Path) -> ParsedLedger:
    if not path.exists():
        raise FileNotFoundError(f"台账文件不存在：{path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    warnings: list[str] = []
    image_map: dict[int, list[LedgerImage]] = defaultdict(list)
    table_path = path
    converted_path: Path | None = None

    if path.suffix.lower() == ".xlsx":
        image_map = extract_xlsx_images(path, work_dir / "images")
        rows = read_xlsx_rows(path)
    elif path.suffix.lower() == ".xls":
        converted_path = try_convert_xls_to_xlsx(path, work_dir)
        if converted_path:
            image_map = extract_xlsx_images(converted_path, work_dir / "images")
            rows = read_xlsx_rows(converted_path)
            table_path = converted_path
        else:
            warnings.append("未能将 .xls 转换为 .xlsx，已生成纯文本日报；如需插入图片，请安装 Microsoft Excel 或改用 .xlsx 台账。")
            rows = read_xls_rows(path)
    else:
        raise ValueError("台账仅支持 .xls 或 .xlsx 文件")

    if len(rows) < 2:
        raise ValueError("台账至少需要两行表头")

    group_headers = fill_merged_headers([clean_cell(value) for value in rows[0]])
    raw_headers = [clean_cell(value) for value in rows[1]]
    headers = [header or group_headers[index] for index, header in enumerate(raw_headers)]
    field_indexes = build_field_indexes(group_headers, headers)
    records = build_records(rows[2:], field_indexes, image_map)
    available_dates: dict[date, int] = {}
    for record in records:
        if record.check_date:
            available_dates[record.check_date] = available_dates.get(record.check_date, 0) + 1

    return ParsedLedger(
        records=records,
        available_dates=dict(sorted(available_dates.items())),
        warnings=warnings,
        image_count=sum(len(images) for images in image_map.values()),
        source_path=table_path,
    )


def fill_merged_headers(headers: list[str]) -> list[str]:
    filled: list[str] = []
    current = ""
    for header in headers:
        if header:
            current = header
        filled.append(current)
    return filled


def build_field_indexes(group_headers: list[str], headers: list[str]) -> dict[str, object]:
    def first_named(*names: str) -> int | None:
        for name in names:
            for index, header in enumerate(headers):
                if header == name:
                    return index
        return None

    required = {
        "serial_no": first_named("编号"),
        "point_level_1": first_named("1级点位"),
        "point_type": first_named("2级点位"),
        "street": first_named("3级点位"),
        "station_name": first_named("4级点位"),
        "issue_text": first_named("具体问题"),
        "indicator_level_1": first_named("1级指标"),
        "indicator_name": first_named("2级指标"),
        "indicator_result": first_named("3级指标"),
        "report_time": first_named("案件上报时间"),
        "created_time": first_named("创建时间"),
    }
    optional_fields = {"report_time", "created_time", "indicator_result"}
    missing = [name for name, index in required.items() if index is None and name not in optional_fields]
    if missing:
        labels = {
            "serial_no": "编号",
            "point_level_1": "1级点位",
            "point_type": "2级点位",
            "street": "3级点位",
            "station_name": "4级点位",
            "issue_text": "具体问题",
            "indicator_level_1": "1级指标",
            "indicator_name": "2级指标",
            "indicator_result": "3级指标",
        }
        raise ValueError("台账缺少必要字段：" + "、".join(labels[name] for name in missing))
    if required["created_time"] is None and required["report_time"] is None:
        raise ValueError("台账缺少必要字段：创建时间或案件上报时间")

    problem_photo_columns = [
        index
        for index, (group, header) in enumerate(zip(group_headers, headers))
        if group in PHOTO_GROUP_NAMES and header.startswith("图片")
    ]
    required["problem_photo_columns"] = problem_photo_columns
    return required


def build_records(
    data_rows: list[list[object]],
    field_indexes: dict[str, object],
    image_map: dict[int, list[LedgerImage]],
) -> list[LedgerRecord]:
    records: list[LedgerRecord] = []
    problem_photo_columns = set(field_indexes.get("problem_photo_columns") or [])

    def value(row: list[object], key: str) -> object:
        index = field_indexes.get(key)
        if index is None:
            return ""
        index = int(index)
        return row[index] if index < len(row) else ""

    for offset, row in enumerate(data_rows, start=3):
        if not any(clean_cell(cell) for cell in row):
            continue
        created_time = parse_datetime_like(value(row, "created_time"))
        report_time = parse_datetime_like(value(row, "report_time"))
        check_date = report_date_from_created_time(created_time or report_time)
        point_type = normalize_station_type(clean_cell(value(row, "point_type")))
        if point_type not in SUPPORTED_TYPES:
            continue
        row_images = sorted(
            [image for image in image_map.get(offset, []) if image.column_index in problem_photo_columns],
            key=lambda image: image.column_index,
        )
        records.append(
            LedgerRecord(
                row_index=offset,
                serial_no=clean_cell(value(row, "serial_no")),
                point_level_1=clean_cell(value(row, "point_level_1")),
                point_type=point_type,
                street=clean_cell(value(row, "street")),
                station_name=clean_cell(value(row, "station_name")),
                issue_text=clean_cell(value(row, "issue_text")),
                indicator_level_1=clean_cell(value(row, "indicator_level_1")),
                indicator_name=clean_cell(value(row, "indicator_name")),
                indicator_result=clean_cell(value(row, "indicator_result")),
                report_time=report_time,
                created_time=created_time,
                check_date=check_date,
                problem_photos=tuple(image.path for image in row_images),
            )
        )
    return records


def read_xlsx_rows(path: Path) -> list[list[object]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheet = workbook.active
    return [[cell.value for cell in row] for row in sheet.iter_rows()]


def read_xls_rows(path: Path) -> list[list[object]]:
    book = xlrd.open_workbook(str(path), formatting_info=False)
    sheet = book.sheet_by_index(0)
    rows: list[list[object]] = []
    for row_index in range(sheet.nrows):
        values: list[object] = []
        for col_index in range(sheet.ncols):
            cell = sheet.cell(row_index, col_index)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    values.append(xlrd.xldate.xldate_as_datetime(cell.value, book.datemode))
                except Exception:
                    values.append(cell.value)
            else:
                values.append(cell.value)
        rows.append(values)
    return rows


def extract_xlsx_images(path: Path, output_dir: Path) -> dict[int, list[LedgerImage]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    result: dict[int, list[LedgerImage]] = defaultdict(list)
    for image_index, image in enumerate(getattr(sheet, "_images", []), start=1):
        anchor = image.anchor
        if not hasattr(anchor, "_from"):
            continue
        row_index = int(anchor._from.row) + 1
        column_index = int(anchor._from.col)
        target = output_dir / f"row{row_index}_col{column_index + 1}_{image_index}.png"
        try:
            raw = image._data()
            write_png_image(raw, target)
        except Exception:
            continue
        result[row_index].append(LedgerImage(row_index=row_index, column_index=column_index, path=target))
    return result


def write_png_image(raw: bytes, target: Path) -> None:
    from io import BytesIO

    with Image.open(BytesIO(raw)) as image:
        if image.mode not in {"RGB", "RGBA"}:
            image = image.convert("RGB")
        image.save(target, format="PNG")


def try_convert_xls_to_xlsx(path: Path, work_dir: Path) -> Path | None:
    target = work_dir / f"{path.stem}_{uuid.uuid4().hex[:8]}.xlsx"
    if try_convert_with_excel(path, target):
        return target
    if try_convert_with_libreoffice(path, target):
        return target
    return None


def try_convert_with_excel(source: Path, target: Path) -> bool:
    try:
        import pythoncom
        import win32com.client
    except Exception:
        return False

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(source.resolve()))
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook.SaveAs(str(target.resolve()), FileFormat=51)
        return target.exists()
    except Exception:
        return False
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def try_convert_with_libreoffice(source: Path, target: Path) -> bool:
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return False
    target.parent.mkdir(parents=True, exist_ok=True)
    import subprocess

    result = subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(target.parent.resolve()),
            str(source.resolve()),
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    generated = target.parent / f"{source.stem}.xlsx"
    if result.returncode != 0 or not generated.exists():
        return False
    if generated.resolve() != target.resolve():
        generated.replace(target)
    return target.exists()
